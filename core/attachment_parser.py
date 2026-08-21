# -*- coding: utf-8 -*-
"""업로드 첨부파일을 프롬프트에 넣을 수 있는 짧은 텍스트 요약으로 변환한다."""
from __future__ import annotations

import csv
import io
from pathlib import Path

MAX_FILES = 5
MAX_TEXT_CHARS = 8000
MAX_ROWS = 30
MAX_COLS = 12


def build_attachment_summary(uploaded_files) -> str | None:
    if not uploaded_files:
        return None

    summaries = []
    for file in uploaded_files[:MAX_FILES]:
        name = getattr(file, "name", "uploaded_file")
        suffix = Path(name).suffix.lower()
        data = file.getvalue()
        try:
            if suffix in {".xlsx", ".xlsm"}:
                body = _summarize_xlsx(data)
            elif suffix == ".csv":
                body = _summarize_csv(data)
            elif suffix == ".pdf":
                body = _summarize_pdf(data)
            elif suffix in {".txt", ".md"}:
                body = _decode_text(data)[:MAX_TEXT_CHARS]
            else:
                body = f"지원하지 않는 파일 형식입니다. 파일명만 참고: {name}"
        except Exception as exc:
            body = f"파일 내용을 읽지 못했습니다: {exc}"
        summaries.append(f"## 첨부파일: {name}\n{body}")

    if len(uploaded_files) > MAX_FILES:
        summaries.append(f"## 생략된 첨부파일\n{len(uploaded_files) - MAX_FILES}개 파일은 요약 한도 때문에 생략됨")

    return "\n\n".join(summaries)


def _decode_text(data: bytes) -> str:
    for enc in ("utf-8-sig", "cp949", "euc-kr", "utf-8"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _summarize_csv(data: bytes) -> str:
    text = _decode_text(data)
    reader = csv.reader(io.StringIO(text))
    rows = []
    for idx, row in enumerate(reader):
        if idx >= MAX_ROWS:
            break
        rows.append(row[:MAX_COLS])
    return _format_rows("CSV 미리보기", rows)


def _summarize_xlsx(data: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames[:3]:
        ws = wb[sheet_name]
        rows = []
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx >= MAX_ROWS:
                break
            values = ["" if v is None else str(v)[:120] for v in row[:MAX_COLS]]
            if any(v.strip() for v in values):
                rows.append(values)
        parts.append(_format_rows(f"시트: {sheet_name}", rows))
    wb.close()
    return "\n\n".join(parts)[:MAX_TEXT_CHARS]


def _summarize_pdf(data: bytes) -> str:
    import fitz

    doc = fitz.open(stream=data, filetype="pdf")
    texts = []
    for page_idx in range(min(5, doc.page_count)):
        page = doc.load_page(page_idx)
        page_text = page.get_text("text").strip()
        if page_text:
            texts.append(page_text[:2000])
    doc.close()
    return "\n\n".join(texts)[:MAX_TEXT_CHARS] or "PDF에서 텍스트를 추출하지 못했습니다."


def _format_rows(title: str, rows: list[list[str]]) -> str:
    if not rows:
        return f"{title}: 읽을 수 있는 행이 없습니다."
    lines = [title]
    for i, row in enumerate(rows, 1):
        lines.append(f"{i}: " + " | ".join(str(v) for v in row))
    return "\n".join(lines)
